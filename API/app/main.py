import io
from typing import Optional, Dict, Any, List
from pathlib import Path
import sys, logging, re, unicodedata
from fastapi.responses import FileResponse, RedirectResponse
from pypdf.errors import PdfReadError, PdfStreamError
from fastapi import FastAPI, UploadFile, File, Form, Header, HTTPException, Query
from fastapi.responses import StreamingResponse, JSONResponse
from pypdf import PdfReader, PdfWriter
from pypdf.generic import NameObject, BooleanObject
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import openpyxl
import os, json
import gspread
from google.oauth2.service_account import Credentials

from .settings import settings
# --- monter l'UI Gradio à la racine ---
import gradio as gr
from .build_ui import build_demo  # adapte l'import si besoin (chemin relatif au repo)

# --- logging global (lisible sur Render aussi) ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger("quitus-api")

# --- helper nom de fichier propre (pour quitus_fullname.pdf) ---
def safe_filename(full_name: str, fallback: str = "document") -> str:
    s = unicodedata.normalize("NFKD", (full_name or "")).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")
    return s or fallback

# --------------------------- App & chemins ---------------------------
app = FastAPI(title="Quitus Filler API")

BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / "templates"          # un seul modèle: quitus.pdf
DATA_DIR = BASE_DIR.parent / "data"
DATA_DIR.mkdir(exist_ok=True)
EXCEL_PATH = DATA_DIR / "students_data.xlsx"



# ------------------------------ Sécurité ----------------------------
def require_api_key(x_api_key: Optional[str] = Header(default=None)) -> None:
    if settings.API_KEY and x_api_key != settings.API_KEY:
        raise HTTPException(status_code=401, detail="Invalid API key")

# --------------------------- Helpers PDF ----------------------------
def get_fields(reader: PdfReader) -> Dict[str, Dict[str, Any]]:
    try:
        return reader.get_fields() or {}
    except Exception:
        return {}

def as_text(v: Any) -> str:
    if isinstance(v, dict):
        v = v.get("/V", "")
    return "" if v is None else str(v).strip()

def extract_all_values(fields: Dict[str, Dict[str, Any]]) -> Dict[str, str]:
    return {name: as_text(obj) for name, obj in fields.items()}

def fill_acroform(base_reader: PdfReader, mapping: Dict[str, str]) -> bytes:
    writer = PdfWriter()
    writer.clone_document_from_reader(base_reader)
    if "/AcroForm" in writer._root_object:
        writer._root_object["/AcroForm"][NameObject("/NeedAppearances")] = BooleanObject(True)
    writer.update_page_form_field_values(writer.pages[0], mapping)
    buf = io.BytesIO(); writer.write(buf); buf.seek(0)
    return buf.read()

def overlay_text(base_reader: PdfReader, lines: List[tuple[str, float, float]]) -> bytes:
    writer = PdfWriter()
    for p in base_reader.pages:
        writer.add_page(p)
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=A4); c.setFont("Helvetica", 12)
    for txt, x, y in lines: c.drawString(x, y, txt)
    c.save(); packet.seek(0)
    overlay_pdf = PdfReader(packet)
    writer.pages[0].merge_page(overlay_pdf.pages[0])
    out = io.BytesIO(); writer.write(out); out.seek(0)
    return out.read()

def get_template_path() -> Path:
    path = TEMPLATES_DIR / "quitus.pdf"   # <— modèle unique
    if not path.exists():
        raise HTTPException(status_code=500, detail="Template not found: quitus.pdf")
    return path

# --------------------------- Helpers Excel --------------------------
def ensure_sheet(wb: openpyxl.Workbook, name: str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)

def read_header(ws) -> List[str]:
    return [(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []

def write_header(ws, header: List[str]) -> None:
    if ws.max_row == 0:
        ws.append(header)
    else:
        for i, h in enumerate(header, start=1):
            ws.cell(row=1, column=i).value = h

def append_row_all_fields(doc_type: str, values: Dict[str, str]) -> None:
    """Log TOUS les champs du PDF source. Feuille selon doc_type."""
    sheet = "Licence" if doc_type == "licence" else "Master"
    wb = openpyxl.load_workbook(EXCEL_PATH) if EXCEL_PATH.exists() else openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = ensure_sheet(wb, sheet)
    header = read_header(ws)
    # ajoute les nouvelles colonnes à la volée
    for k in values.keys():
        if k not in header:
            header.append(k)
    write_header(ws, header)
    ws.append([values.get(col, "") for col in header])
    wb.save(EXCEL_PATH)

# -------- Google Sheets (persistant) --------
_GS_CLIENT = None
_GS_SHEET = None

GS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    # Ajoute Drive si tu veux créer le fichier automatiquement
    "https://www.googleapis.com/auth/drive",
]

def _gs_client():
    global _GS_CLIENT
    if _GS_CLIENT is None:
        sa_json = settings.GCP_SA_JSON
        if not sa_json:
            raise RuntimeError("GCP_SA_JSON manquant dans les variables d'environnement")
        creds = Credentials.from_service_account_info(json.loads(sa_json), scopes=GS_SCOPES)
        _GS_CLIENT = gspread.authorize(creds)
    return _GS_CLIENT

def _gs_sheet():
    global _GS_SHEET
    if _GS_SHEET is not None:
        return _GS_SHEET

    gc = _gs_client()
    sheet_id = settings.GSHEET_ID
    sheet_name = settings.GSHEET_NAME
    allow_create = bool(settings.GSHEET_CREATE)

    try:
        if sheet_id:
            sh = gc.open_by_key(sheet_id)
        else:
            try:
                sh = gc.open(sheet_name)
            except gspread.SpreadsheetNotFound:
                if not allow_create:
                    raise
                sh = gc.create(sheet_name)
    except Exception as e:
        raise RuntimeError(f"Impossible d'ouvrir le Google Sheet: {e}")

    _GS_SHEET = sh
    return _GS_SHEET


def _gs_ensure_worksheet(sh, title: str):
    try:
        return sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        return sh.add_worksheet(title=title, rows=100, cols=26)

def _gs_read_header(ws) -> list[str]:
    values = ws.row_values(1) if ws.row_count >= 1 else []
    return [h.strip() for h in values] if values else []

def _gs_write_header(ws, header: list[str]):
    """Écrit l'entête sur la ligne 1, en étendant au besoin."""
    if not header:
        return
    # (facultatif) s'assurer d'avoir assez de colonnes
    needed = len(header) - ws.col_count
    if needed > 0:
        ws.add_cols(needed)

    # Version la plus simple et sûre : commencer à A1 et laisser gspread étendre
    ws.update("A1", [header])

def append_row_all_fields_sheets(doc_type: str, values: dict[str, str]) -> None:
    """
    Persiste TOUS les champs dans Google Sheets.
    - Feuille 'Licence' ou 'Master'
    - Ajoute dynamiquement les colonnes manquantes
    - Aligne la ligne sur l'entête
    """
    sh = _gs_sheet()
    ws = _gs_ensure_worksheet(sh, "Licence" if doc_type == "licence" else "Master")

    header = _gs_read_header(ws)
    if not header:
        header = list(values.keys())
        _gs_write_header(ws, header)
    else:
        # ajoute les nouvelles colonnes à la fin
        changed = False
        for k in values.keys():
            if k not in header:
                header.append(k); changed = True
        if changed:
            _gs_write_header(ws, header)

    row = [values.get(col, "") for col in header]
    ws.append_row(row, value_input_option="USER_ENTERED")

# ------------------------------- Routes -----------------------------
# Redirection claire de la racine vers l'UI (évite le //)
@app.get("/", include_in_schema=False)
def root_redirect():
    return RedirectResponse(url="/app", status_code=307)
@app.get("/manifest.json")
def manifest():
    return JSONResponse({
        "name": "Quitus Filler",
        "short_name": "Quitus",
        "start_url": "/",
        "display": "standalone",
        "icons": []
    })
@app.post("/process")
async def process_quitus(
    source_pdf: UploadFile = File(...),
    doc_type: Optional[str] = Form(None),     # accept form
    doc_type_q: Optional[str] = Query(None),  # ou query
    x_api_key: Optional[str] = Header(default=None),
):
    require_api_key(x_api_key)
    dt = (doc_type or doc_type_q or "licence").lower()
    if dt not in {"licence", "master"}:
        raise HTTPException(status_code=400, detail="doc_type must be 'licence' or 'master'")

    # 1) lire les octets UNE fois + valider (%PDF)
    data = await source_pdf.read()
    if not data or len(data) < 5 or not data.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Invalid or empty PDF (missing %PDF header)")
    try:
        src_reader = PdfReader(io.BytesIO(data), strict=False)
    except (PdfReadError, PdfStreamError) as e:
        raise HTTPException(status_code=400, detail=f"Unreadable PDF: {e}")

    # 2) extraire champs
    all_values = extract_all_values(get_fields(src_reader))
    nom = all_values.get("student_nom", "")
    prenom = all_values.get("student_prenom", "")
    cin = all_values.get("cin", "")
    filiere_lic = all_values.get("filiere_lic", "")
    filiere_master = all_values.get("filiere_master", "")
    full_name = f"{nom} {prenom}".strip()
    filiere_finale = filiere_lic if dt == "licence" else (filiere_master or filiere_lic)

   
    # 3) Persistance
    mode = settings.EXCEL_MODE.lower()
    if mode == "gsheets":
        append_row_all_fields_sheets(dt, all_values)
    else:
        append_row_all_fields(dt, all_values)
   
    # 4) remplir modèle
    with open(get_template_path(), "rb") as fh:
        q_reader = PdfReader(io.BytesIO(fh.read()), strict=False)
    mapping = {"student_nom": full_name, "cin": cin, "filiere_lic": filiere_finale}
    q_fields = get_fields(q_reader)
    if q_fields:
        pdf_out = fill_acroform(q_reader, mapping)
    else:
        lines = [(full_name, 120, 690), (cin, 160, 665), (filiere_finale, 160, 640)]
        pdf_out = overlay_text(q_reader, lines)

    # 5) nommage quitus_<fullname>.pdf
    slug = safe_filename(full_name, fallback=dt)
    filename = f"quitus_{slug}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_out),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/health")
def health():
    return {
        "ok": True,
        "excel_exists": EXCEL_PATH.exists(),
        "excel_path": str(EXCEL_PATH),
        "template_exists": (TEMPLATES_DIR / "quitus.pdf").exists(),
    }

@app.get("/download/excel")
def download_excel(
    sheet: Optional[str] = Query(None, description="Licence | Master (vide = les deux)"),
    x_api_key: Optional[str] = Header(default=None),
):
    require_api_key(x_api_key)

    mode = settings.EXCEL_MODE.lower()
    if mode == "gsheets":
        # Export Google Sheets -> XLSX en mémoire
        sh = _gs_sheet()

        # Prépare le classeur de sortie
        wb = openpyxl.Workbook()
        # supprime la feuille par défaut si présente
        if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 and not wb.active["A1"].value:
            wb.remove(wb.active)

        def add_ws(title: str):
            try:
                ws_g = sh.worksheet(title)
            except gspread.exceptions.WorksheetNotFound:
                return False
            values = ws_g.get_all_values()  # liste de listes
            ws_x = wb.create_sheet(title=title)
            for row in values:
                ws_x.append(row)
            # fige l'entête si elle existe
            if values:
                ws_x.freeze_panes = "A2"
            return True

        added = False
        if sheet:
            added = add_ws(sheet)
        else:
            # essaie dans cet ordre
            for t in ("Licence", "Master"):
                added = add_ws(t) or added

        if not added:
            raise HTTPException(status_code=404, detail="No data in Google Sheets yet")

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="students_data.xlsx"'}
        )

    # ---- mode local (fichier sur disque) ----
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=404, detail="No Excel yet")
    return FileResponse(
        EXCEL_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="students_data.xlsx",
    )

demo = build_demo(default_api_url="/process")  # même service
app = gr.mount_gradio_app(app, demo, path="/app") # l'UI sert "/" ; l'API reste dispo (ex: /process, /health, /docs)
